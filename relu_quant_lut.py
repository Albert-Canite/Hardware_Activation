import numpy as np
import torch
import torch.nn as nn
from openpyxl import load_workbook


class QuantReLU8bit(nn.Module):
    """ReLU with 8-bit quantization on both input and output over [-1, 1]."""

    def __init__(self, min_val: float = -1.0, max_val: float = 1.0, levels: int = 256):
        super().__init__()
        self.min_val = float(min_val)
        self.max_val = float(max_val)
        self.levels = int(levels)
        self.step = (self.max_val - self.min_val) / (self.levels - 1)

    def quantize(self, x: torch.Tensor) -> torch.Tensor:
        x = torch.clamp(x, self.min_val, self.max_val)
        x = torch.round((x - self.min_val) / self.step) * self.step + self.min_val
        return torch.clamp(x, self.min_val, self.max_val)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        x_q = self.quantize(x)
        y = torch.relu(x_q)
        return self.quantize(y)


class LUTReLU8bit(nn.Module):
    """Lookup-table based ReLU using a precomputed 256-entry mapping over [-1, 1]."""

    def __init__(self, lut_y: np.ndarray, min_val: float = -1.0, max_val: float = 1.0):
        super().__init__()
        if len(lut_y) != 256:
            raise ValueError(f"LUT length must be 256, got {len(lut_y)}")
        self.min_val = float(min_val)
        self.max_val = float(max_val)
        self.levels = 256
        self.step = (self.max_val - self.min_val) / (self.levels - 1)
        self.register_buffer("lut", torch.tensor(lut_y, dtype=torch.float32))

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        x = torch.clamp(x, self.min_val, self.max_val)
        idx = torch.round((x - self.min_val) / self.step).long()
        idx = torch.clamp(idx, 0, self.levels - 1)
        return self.lut[idx]


def replace_relu_with_quantrelu(module: nn.Module) -> None:
    for name, child in module.named_children():
        if isinstance(child, nn.ReLU):
            setattr(module, name, QuantReLU8bit())
        else:
            replace_relu_with_quantrelu(child)


def replace_quantrelu_with_lutrelu(module: nn.Module, lut_y_256: np.ndarray) -> None:
    for name, child in module.named_children():
        if isinstance(child, QuantReLU8bit):
            setattr(module, name, LUTReLU8bit(lut_y_256))
        else:
            replace_quantrelu_with_lutrelu(child, lut_y_256)


def load_lut_from_excel(xlsx_path: str) -> tuple[np.ndarray, np.ndarray]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    xs, ys = [], []
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        x, y = row
        if x is None or y is None:
            continue
        try:
            xs.append(float(x))
            ys.append(float(y))
        except (TypeError, ValueError):
            continue

    if len(xs) < 2:
        raise ValueError("LUT_ReLU.xlsx must contain at least 2 numeric rows with [input, output].")

    xs = np.asarray(xs, dtype=np.float32)
    ys = np.asarray(ys, dtype=np.float32)

    order = np.argsort(xs)
    return xs[order], ys[order]


def interpolate_lut_to_8bit(xs: np.ndarray, ys: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    x_256 = np.linspace(-1.0, 1.0, 256, dtype=np.float32)
    y_256 = np.interp(x_256, xs, ys).astype(np.float32)
    y_256 = np.clip(y_256, -1.0, 1.0)
    return x_256, y_256


def save_lut_8bit_csv(x_256: np.ndarray, y_256: np.ndarray, output_csv: str) -> None:
    data = np.stack([x_256, y_256], axis=1)
    np.savetxt(output_csv, data, delimiter=",", header="input_8bit,output_8bit", comments="")
