import argparse
import copy
import faulthandler
import os
import platform
import random
import subprocess
import sys
from dataclasses import dataclass
from typing import Callable, Dict, Tuple

import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
from torchvision import datasets, transforms

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise ImportError("openpyxl is required to read LUT_ReLU.xlsx.") from exc


QMIN, QMAX, QLEVELS = -1.0, 1.0, 256


def set_seed(seed: int = 42) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)


def select_device(device_arg: str = "auto", require_gpu: bool = False) -> torch.device:
    """
    Robust device selector.
    - auto: prefer CUDA if available AND a subprocess smoke test passes.
    - cuda: require CUDA smoke test pass, else fallback to CPU with warning.
    - cpu: force CPU.
    """
    if device_arg == "cpu":
        if require_gpu:
            raise RuntimeError("GPU is required (--require-gpu) but --device cpu was provided.")
        return torch.device("cpu")

    if not torch.cuda.is_available():
        if require_gpu or device_arg == "cuda":
            raise RuntimeError("CUDA is not available, but GPU execution is required.")
        return torch.device("cpu")

    # 子进程做 CUDA 烟雾测试，避免主进程直接崩溃（0xC0000005）
    probe_code = (
        "import torch; "
        "assert torch.cuda.is_available(); "
        "x=torch.randn(64,64,device='cuda'); "
        "y=torch.relu(x); "
        "z=(y@y.T).sum(); "
        "z.item(); "
        "print('CUDA_PROBE_OK')"
    )
    try:
        p = subprocess.run(
            [sys.executable, "-c", probe_code],
            capture_output=True,
            text=True,
            timeout=20,
        )
        if p.returncode == 0 and "CUDA_PROBE_OK" in p.stdout:
            return torch.device("cuda")
        if require_gpu or device_arg == "cuda":
            raise RuntimeError(
                "CUDA probe failed (returncode={}): {} {}".format(
                    p.returncode, p.stdout.strip(), p.stderr.strip()
                )
            )
        print("[WARN] CUDA probe failed in auto mode. Falling back to CPU.")
        return torch.device("cpu")
    except Exception as exc:
        if require_gpu or device_arg == "cuda":
            raise RuntimeError(f"CUDA probe exception: {exc}")
        print(f"[WARN] CUDA probe exception: {exc}. Falling back to CPU.")
        return torch.device("cpu")


def quantize_8bit(x: torch.Tensor, qmin: float = QMIN, qmax: float = QMAX) -> torch.Tensor:
    x = torch.clamp(x, qmin, qmax)
    x_scaled = (x - qmin) / (qmax - qmin) * (QLEVELS - 1)
    x_rounded = torch.round(x_scaled)
    return x_rounded / (QLEVELS - 1) * (qmax - qmin) + qmin


class QuantizedReLU(nn.Module):
    """Input/output are quantized to 8-bit in [-1, 1], activation is standard ReLU."""

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        x_q = quantize_8bit(x)
        y = torch.relu(x_q)
        return quantize_8bit(y)


class LUTReLU(nn.Module):
    """Input/output quantized to 8-bit in [-1, 1], activation by LUT lookup."""

    def __init__(self, lut_x: np.ndarray, lut_y: np.ndarray):
        super().__init__()
        assert lut_x.shape[0] == QLEVELS and lut_y.shape[0] == QLEVELS
        self.register_buffer("lut_x", torch.tensor(lut_x, dtype=torch.float32))
        self.register_buffer("lut_y", torch.tensor(lut_y, dtype=torch.float32))

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        x_q = quantize_8bit(x)
        idx = torch.round((x_q - QMIN) / (QMAX - QMIN) * (QLEVELS - 1)).long()
        idx = torch.clamp(idx, 0, QLEVELS - 1)
        y = self.lut_y[idx]
        return quantize_8bit(y)


CFG_VGG11 = [64, "M", 128, "M", 256, 256, "M", 512, 512, "M", 512, 512, "M"]


class VGG11Small(nn.Module):
    def __init__(self, num_classes: int, in_channels: int, activation_factory: Callable[[], nn.Module]):
        super().__init__()
        self.features = self._make_layers(CFG_VGG11, in_channels, activation_factory)
        self.classifier = nn.Sequential(
            nn.Flatten(),
            nn.Linear(512, 512),
            activation_factory(),
            nn.Dropout(0.3),
            nn.Linear(512, num_classes),
        )

    @staticmethod
    def _make_layers(cfg, in_channels, activation_factory):
        layers = []
        for x in cfg:
            if x == "M":
                layers.append(nn.MaxPool2d(kernel_size=2, stride=2))
            else:
                layers.extend([
                    nn.Conv2d(in_channels, x, kernel_size=3, padding=1, bias=False),
                    nn.BatchNorm2d(x),
                    activation_factory(),
                ])
                in_channels = x
        return nn.Sequential(*layers)

    def forward(self, x):
        x = self.features(x)
        return self.classifier(x)


@dataclass
class DatasetSpec:
    name: str
    num_classes: int
    in_channels: int


def get_dataloaders(dataset_name: str, data_root: str, batch_size: int, num_workers: int, pin_memory: bool):
    if dataset_name.lower() == "mnist":
        transform = transforms.Compose([
            transforms.Resize((32, 32)),
            transforms.ToTensor(),
            transforms.Normalize((0.1307,), (0.3081,)),
        ])
        train_set = datasets.MNIST(data_root, train=True, download=True, transform=transform)
        test_set = datasets.MNIST(data_root, train=False, download=True, transform=transform)
        spec = DatasetSpec("MNIST", 10, 1)
    elif dataset_name.lower() == "cifar10":
        transform = transforms.Compose([
            transforms.ToTensor(),
            transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2470, 0.2435, 0.2616)),
        ])
        train_set = datasets.CIFAR10(data_root, train=True, download=True, transform=transform)
        test_set = datasets.CIFAR10(data_root, train=False, download=True, transform=transform)
        spec = DatasetSpec("CIFAR10", 10, 3)
    else:
        raise ValueError("dataset_name must be one of: mnist, cifar10")

    train_loader = DataLoader(
        train_set,
        batch_size=batch_size,
        shuffle=True,
        num_workers=num_workers,
        pin_memory=pin_memory,
        persistent_workers=num_workers > 0,
    )
    test_loader = DataLoader(
        test_set,
        batch_size=batch_size,
        shuffle=False,
        num_workers=num_workers,
        pin_memory=pin_memory,
        persistent_workers=num_workers > 0,
    )
    return spec, train_loader, test_loader


def train_one_epoch(model, loader, criterion, optimizer, device, max_batches=None):
    model.train()
    running_loss = 0.0
    total = 0
    correct = 0
    for batch_idx, (x, y) in enumerate(loader):
        if max_batches is not None and batch_idx >= max_batches:
            break
        x, y = x.to(device, non_blocking=True), y.to(device, non_blocking=True)
        optimizer.zero_grad(set_to_none=True)
        logits = model(x)
        loss = criterion(logits, y)
        loss.backward()
        optimizer.step()

        running_loss += loss.item() * x.size(0)
        pred = logits.argmax(dim=1)
        total += y.size(0)
        correct += (pred == y).sum().item()

    return running_loss / max(total, 1), correct / max(total, 1)


@torch.no_grad()
def evaluate(model, loader, device, max_batches=None):
    model.eval()
    total = 0
    correct = 0
    for batch_idx, (x, y) in enumerate(loader):
        if max_batches is not None and batch_idx >= max_batches:
            break
        x, y = x.to(device, non_blocking=True), y.to(device, non_blocking=True)
        logits = model(x)
        pred = logits.argmax(dim=1)
        total += y.size(0)
        correct += (pred == y).sum().item()
    return correct / max(total, 1)


def build_interpolated_lut(lut_xlsx: str, output_csv: str) -> Tuple[np.ndarray, np.ndarray]:
    # Use openpyxl directly to avoid pandas engine/native-stack instability on some Windows setups.
    wb = load_workbook(lut_xlsx, read_only=True, data_only=True)
    ws = wb.active
    x_vals = []
    y_vals = []
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row is None:
            continue
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        try:
            xa = float(a)
            yb = float(b)
            if np.isfinite(xa) and np.isfinite(yb):
                x_vals.append(xa)
                y_vals.append(yb)
        except (TypeError, ValueError):
            continue
    wb.close()

    x_src = np.asarray(x_vals, dtype=np.float32)
    y_src = np.asarray(y_vals, dtype=np.float32)

    if x_src.size < 2:
        raise ValueError("LUT_ReLU.xlsx does not contain enough numeric (input, output) pairs")

    order = np.argsort(x_src)
    x_src, y_src = x_src[order], y_src[order]

    x_8bit = np.linspace(QMIN, QMAX, QLEVELS, dtype=np.float32)
    y_8bit = np.interp(x_8bit, x_src, y_src).astype(np.float32)
    y_8bit = np.clip(y_8bit, QMIN, QMAX)
    y_8bit = np.round((y_8bit - QMIN) / (QMAX - QMIN) * (QLEVELS - 1)) / (QLEVELS - 1) * (QMAX - QMIN) + QMIN

    os.makedirs(os.path.dirname(output_csv), exist_ok=True)
    np.savetxt(output_csv, np.stack([x_8bit, y_8bit], axis=1), delimiter=",", header="input,output", comments="")
    return x_8bit, y_8bit


def replace_quant_relu_with_lut(module: nn.Module, lut_x: np.ndarray, lut_y: np.ndarray):
    for name, child in module.named_children():
        if isinstance(child, QuantizedReLU):
            setattr(module, name, LUTReLU(lut_x, lut_y))
        else:
            replace_quant_relu_with_lut(child, lut_x, lut_y)


def run_for_dataset(args, dataset_name: str, lut_x: np.ndarray, lut_y: np.ndarray, device: torch.device) -> Dict[str, float]:
    print(f"[INFO] Preparing dataloaders for {dataset_name}...")
    spec, train_loader, test_loader = get_dataloaders(
        dataset_name,
        args.data_root,
        args.batch_size,
        args.num_workers,
        pin_memory=(device.type == "cuda" and not args.disable_pin_memory),
    )

    print(f"[INFO] Building model for {spec.name}...")
    model = VGG11Small(spec.num_classes, spec.in_channels, activation_factory=QuantizedReLU).to(device)
    criterion = nn.CrossEntropyLoss()
    optimizer = optim.Adam(model.parameters(), lr=args.lr)

    print(f"\n===== Training {spec.name} (8-bit quantized ReLU) =====")
    for epoch in range(args.epochs):
        tr_loss, tr_acc = train_one_epoch(
            model, train_loader, criterion, optimizer, device, max_batches=args.max_train_batches
        )
        te_acc = evaluate(model, test_loader, device, max_batches=args.max_test_batches)
        print(f"Epoch {epoch+1}/{args.epochs}: loss={tr_loss:.4f}, train_acc={tr_acc:.4f}, test_acc={te_acc:.4f}")

    os.makedirs(args.output_dir, exist_ok=True)
    ckpt_path = os.path.join(args.output_dir, f"vgg11_{spec.name.lower()}_quant_relu.pth")
    torch.save(model.state_dict(), ckpt_path)
    print(f"Saved model: {ckpt_path}")

    acc_standard = evaluate(model, test_loader, device, max_batches=args.max_test_batches)

    lut_model = VGG11Small(spec.num_classes, spec.in_channels, activation_factory=QuantizedReLU)
    lut_model.load_state_dict(copy.deepcopy(model.state_dict()))
    replace_quant_relu_with_lut(lut_model, lut_x, lut_y)
    lut_model.to(device)
    acc_lut = evaluate(lut_model, test_loader, device, max_batches=args.max_test_batches)

    return {
        "standard_relu_acc": acc_standard,
        "lut_relu_acc": acc_lut,
        "checkpoint": ckpt_path,
    }


def parse_args():
    parser = argparse.ArgumentParser(description="Train/evaluate VGG-11 on MNIST + CIFAR10 with quantized activations and LUT-ReLU replacement")
    parser.add_argument("--data-root", type=str, default="./data")
    parser.add_argument("--output-dir", type=str, default="./outputs")
    parser.add_argument("--lut-file", type=str, default="./LUT_ReLU.xlsx")
    parser.add_argument("--epochs", type=int, default=10)
    parser.add_argument("--batch-size", type=int, default=32)
    parser.add_argument("--lr", type=float, default=1e-3)
    parser.add_argument("--num-workers", type=int, default=0)
    parser.add_argument("--device", type=str, default="auto", choices=["auto", "cpu", "cuda"])
    parser.add_argument("--require-gpu", action="store_true", help="Fail fast if GPU is not usable")
    parser.add_argument("--disable-pin-memory", action="store_true", help="Disable DataLoader pin_memory (can help on Windows GPU issues)")
    parser.add_argument("--gpu-safe-mode", dest="gpu_safe_mode", action="store_true", help="Use safer CUDA runtime settings")
    parser.add_argument("--no-gpu-safe-mode", dest="gpu_safe_mode", action="store_false", help="Disable safer CUDA runtime settings")
    parser.set_defaults(gpu_safe_mode=True)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--max-train-batches", type=int, default=None, help="Optional debug cap per epoch")
    parser.add_argument("--max-test-batches", type=int, default=None, help="Optional debug cap for eval")
    parser.add_argument("--_worker", action="store_true", help=argparse.SUPPRESS)
    return parser.parse_args()




def launch_worker_process() -> int:
    """Run the heavy training flow in a subprocess to avoid hard IDE crash on native CUDA faults."""
    cmd = [sys.executable, os.path.abspath(__file__), "--_worker"] + sys.argv[1:]
    proc = subprocess.Popen(cmd)
    proc.wait()
    code = proc.returncode
    # Windows access violation
    if code == 3221225477 or code == -1073741819:
        print("\n[ERROR] Worker crashed with 0xC0000005 (native access violation).")
        print("[ERROR] This indicates a low-level CUDA/driver/runtime crash, not Python logic.")
        print("[ERROR] Crash is likely outside model code (driver/cuDNN/OS interaction).")
        return 1
    return 0 if code == 0 else code

def main():
    faulthandler.enable(all_threads=True)
    args = parse_args()

    if not args._worker:
        rc = launch_worker_process()
        raise SystemExit(rc)

    # Windows + CUDA + 多进程 DataLoader 在部分 PyTorch 2.2.x 环境中容易触发原生层崩溃
    # （表现为 0xC0000005 access violation）。默认强制单进程读取更稳。
    if platform.system().lower() == "windows" and args.num_workers > 0:
        print("[WARN] Windows detected: forcing num_workers=0 for stability.")
        args.num_workers = 0
    if platform.system().lower() == "windows":
        # Windows 环境默认关闭 pin_memory，避免部分驱动/环境在 page-locked 内存路径崩溃
        args.disable_pin_memory = True

    print("[STAGE] set_seed")
    set_seed(args.seed)
    print("[STAGE] select_device")
    device = select_device(args.device, require_gpu=args.require_gpu)
    print(f"Using device: {device}")

    if device.type == "cuda":
        print("[STAGE] configure_cuda_flags")
        # 关闭 benchmark 可避免部分机器上 cudnn 自动算法选择触发不稳定问题
        torch.backends.cudnn.benchmark = False
        torch.backends.cudnn.deterministic = True
        if args.gpu_safe_mode:
            # 很多 0xC0000005 发生在 cudnn 内核路径，safe mode 下禁用 cudnn 但保留 CUDA 计算
            torch.backends.cudnn.enabled = False
            torch.backends.cuda.matmul.allow_tf32 = False
            print("[INFO] GPU safe mode enabled: cudnn disabled, tf32 disabled.")
        # 注意：不再调用 torch.cuda.set_per_process_memory_fraction。
        # 该调用在部分 Windows + 驱动组合中会直接触发 native 崩溃（无 Python 异常）。

    print("[STAGE] build_lut")
    lut_csv_path = os.path.join(args.output_dir, "relu_lut_8bit.csv")
    lut_x, lut_y = build_interpolated_lut(args.lut_file, lut_csv_path)
    print(f"Saved interpolated 8-bit LUT: {lut_csv_path}")

    mnist_result = run_for_dataset(args, "mnist", lut_x, lut_y, device)
    cifar_result = run_for_dataset(args, "cifar10", lut_x, lut_y, device)

    print("\n===== Final Accuracy Comparison =====")
    print(f"MNIST  - standard ReLU accuracy: {mnist_result['standard_relu_acc']:.4f}")
    print(f"MNIST  - LUT ReLU accuracy:      {mnist_result['lut_relu_acc']:.4f}")
    print(f"CIFAR10- standard ReLU accuracy: {cifar_result['standard_relu_acc']:.4f}")
    print(f"CIFAR10- LUT ReLU accuracy:      {cifar_result['lut_relu_acc']:.4f}")


if __name__ == "__main__":
    main()
